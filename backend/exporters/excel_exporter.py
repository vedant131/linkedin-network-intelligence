"""
Excel exporter — generates a two-sheet .xlsx report using OpenPyXL.
Sheet 1: "My Network"   — full enriched data with auto-filter + color coding
Sheet 2: "Insights"     — pivot-style summary stats
"""
import io
from collections import Counter
import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

# ── Category colour palette ────────────────────────────────────────────────────
CATEGORY_COLOURS = {
    "Software Engineer":      "DBEAFE",  # light blue
    "Data Scientist":         "EDE9FE",  # light violet
    "Recruiter/HR":           "DCFCE7",  # light green
    "Founder/Entrepreneur":   "FEF3C7",  # light amber
    "Student":                "FEE2E2",  # light red
    "Marketing/Sales":        "FCE7F3",  # light pink
    "Other":                  "F1F5F9",  # light slate
}

HEADER_FILL  = PatternFill("solid", fgColor="0F172A")
HEADER_FONT  = Font(bold=True, color="E2E8F0", size=10)
BODY_FONT    = Font(size=9)
BORDER_SIDE  = Side(style="thin", color="E2E8F0")
THIN_BORDER  = Border(bottom=BORDER_SIDE)


def _col_width(df: pd.DataFrame, col: str, min_w: int = 10, max_w: int = 40) -> int:
    header_len = len(str(col)) + 2
    if col in df.columns:
        data_len = df[col].astype(str).str.len().quantile(0.9)
    else:
        data_len = header_len
    return int(min(max(max(header_len, data_len), min_w), max_w))


def build_excel(df: pd.DataFrame) -> bytes:
    """Build and return an Excel workbook as bytes."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: My Network ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "My Network"

    columns = [
        ("Full Name",       "full_name"),
        ("Job Title",       "position_clean"),
        ("Category",        "category"),
        ("Seniority",       "seniority"),
        ("Domain",          "domain"),
        ("Company",         "company_clean"),
        ("Tags",            "tags"),
        ("Email",           "email"),
        ("Connected On",    "connected_on"),
        ("Score",           "score"),
    ]

    # Header row
    for c_idx, (header, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Data rows
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        cat = getattr(row, "category", "Other")
        fill_colour = CATEGORY_COLOURS.get(cat, "F1F5F9")
        row_fill = PatternFill("solid", fgColor=fill_colour)

        for c_idx, (_, field) in enumerate(columns, 1):
            val = getattr(row, field, "")
            if isinstance(val, list):
                val = ", ".join(val)
            if field == "score":
                val = round(float(val), 2) if val else 0.0
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = row_fill
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    # Column widths
    for c_idx, (header, field) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = _col_width(df, field)

    # ── Sheet 2: Insights ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Insights")

    def _write_section(ws, title: str, data: list[tuple], start_row: int, start_col: int) -> int:
        """Write a title + key-value table block, return next free row."""
        title_cell = ws.cell(row=start_row, column=start_col, value=title)
        title_cell.font = Font(bold=True, size=11, color="0F172A")
        title_cell.fill = PatternFill("solid", fgColor="BFDBFE")
        ws.cell(row=start_row, column=start_col + 1).fill = PatternFill("solid", fgColor="BFDBFE")

        for i, (k, v) in enumerate(data, 1):
            ws.cell(row=start_row + i, column=start_col, value=str(k)).font = Font(size=9, bold=True)
            ws.cell(row=start_row + i, column=start_col + 1, value=v).font   = Font(size=9)

        return start_row + len(data) + 2

    # Category breakdown
    cat_counts = df["category"].value_counts().items()
    next_row = _write_section(ws2, "Connections by Category", list(cat_counts), 1, 1)

    # Seniority breakdown
    sen_counts = df["seniority"].value_counts().items()
    next_row = _write_section(ws2, "Connections by Seniority", list(sen_counts), next_row, 1)

    # Top companies
    top_companies = df["company_clean"].value_counts().head(10).items()
    _write_section(ws2, "Top Companies", list(top_companies), 1, 4)

    # Domain breakdown
    dom_counts = df["domain"].value_counts().head(10).items()
    _write_section(ws2, "Top Domains", list(dom_counts), 1, 7)

    # Column widths for Insights
    for col in [1, 2, 4, 5, 7, 8]:
        ws2.column_dimensions[get_column_letter(col)].width = 22

    # Summary row at top of Insights
    ws2.insert_rows(1)
    summary_text = (
        f"Total: {len(df)}  |  "
        f"Recruiters: {(df['category']=='Recruiter/HR').sum()}  |  "
        f"Engineers: {(df['category']=='Software Engineer').sum()}  |  "
        f"High Value: {df['tags'].apply(lambda t: 'High Value Connection' in t).sum()}"
    )
    ws2["A1"] = summary_text
    ws2["A1"].font = Font(bold=True, size=10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
